import openpyxl
from openpyxl.workbook.defined_name import DefinedName
import os

def create_mock_invoice(path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Invoice"
    
    # Dane Nagłówka
    ws['B2'] = "2026-01-20" # date_issue
    ws['B3'] = "2025-10-FV01_TEST" # note_title
    
    # Dane Nabywcy (Jeden blok!)
    ws['A5'] = "BNP PARIBAS Asset Management\nSPÓŁKA AKCYJNA UPROSZCZONA Oddział w Polsce\nul. Wronia 31, 00-846 Warszawa\nNIP: 5273128575"
    
    # Tabela cz. 1 (EUR)
    ws['A9'] = "Lp"
    ws['B9'] = "Nazwa usługi"
    ws['C9'] = "jednostka miary"
    ws['D9'] = "ilość"
    ws['E9'] = "cena jednostkowa netto eur"
    ws['F9'] = "wartość sprzedaży netto eur"
    ws['G9'] = "wartość sprzedaży netto w PLN"
    ws['H9'] = "Stawka %"
    ws['I9'] = "Kwota w PLN"
    
    ws['A10'] = 1 # entries_array punkt startowy
    ws['B10'] = "Podnajem powierzchni"
    ws['C10'] = "m2"
    ws['D10'] = 332
    ws['E10'] = 20.34
    ws['F10'] = 6752.88
    ws['G10'] = 28000
    ws['H10'] = 23
    ws['I10'] = 6440
    
    # Tabela cz. 2 (Tylko PLN) - linia pod linią
    ws['A12'] = "Lp"
    ws['B12'] = "Nazwa usługi"
    ws['C12'] = "jednostka miary"
    ws['D12'] = "ilość"
    ws['E12'] = "cena jednostkowa netto PLN"
    ws['F12'] = "" # pusta kolumna jak w Twoim opisie
    ws['G12'] = "wartość sprzedaży netto w PLN"
    ws['H12'] = "Stawka %"
    ws['I12'] = "Kwota w PLN"

    ws['A13'] = 2
    ws['B13'] = "Oplata infrastruktura"
    ws['C13'] = "szt"
    ws['D13'] = 1
    ws['E13'] = 11222
    ws['G13'] = 11222
    ws['H13'] = 23
    ws['I13'] = 2581
    
    ws['A15'] = "Suma" # Stop

    # Named Ranges
    from openpyxl.workbook.defined_name import DefinedName
    wb.defined_names.add(DefinedName("date_issue", attr_text='Invoice!$B$2'))
    wb.defined_names.add(DefinedName("note_title", attr_text='Invoice!$B$3'))
    wb.defined_names.add(DefinedName("recipient_info", attr_text='Invoice!$A$5'))
    wb.defined_names.add(DefinedName("entries_array", attr_text='Invoice!$A$10')) # Start danych

    os.makedirs(os.path.dirname(path), exist_ok=True)
    wb.save(path)
    print(f"Utworzono atrapę faktury: {path}")

if __name__ == "__main__":
    create_mock_invoice(r"c:\Users\Admin\Desktop\Dudek\Python\Projekty\KSeF-Python-Client-V9_Kopia_Toola_z_modulem_Wiz_Po_Poprawkach_1\storage\test_mock_invoice.xlsx")
