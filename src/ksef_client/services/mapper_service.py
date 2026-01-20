import os
import json
import pandas as pd
from lxml import etree
from datetime import datetime
import openpyxl
import re

class MapperService:
    def __init__(self, config_obj):
        self.cfg = config_obj
        self.ns = {
            None: "http://crd.gov.pl/wzor/2025/06/25/13775/",
            "xsi": "http://www.w3.org/2001/XMLSchema-instance",
            "xsd": "http://www.w3.org/2001/XMLSchema"
        }

    def load_mapping_config(self, mapping_name):
        config_path = self.cfg.config_dir / 'mappings' / f"{mapping_name}.json"
        if not os.path.exists(config_path):
            raise FileNotFoundError(f"Mapping config not found: {config_path}")
        with open(config_path, 'r', encoding='utf-8') as f:
            return json.load(f)

    def get_value_from_excel(self, wb, sheet, mapping_rule):
        source = mapping_rule.get("source")
        val = mapping_rule.get("value")

        if source == "CONSTANT":
            return val
        
        if source == "NAME":
            if val in wb.defined_names:
                d_name = wb.defined_names[val]
                dest = list(d_name.destinations)[0] 
                worksheet = wb[dest[0]]
                return worksheet[dest[1]].value
            return None

        if source == "ADRES":
            return sheet[val].value

        return None

    def _parse_recipient_info(self, raw_text):
        """Rozbija multiline text na NIP, Nazwę i Adres."""
        import re
        lines = [line.strip() for line in raw_text.split('\n') if line.strip()]
        result = {"NIP": "", "Nazwa": "", "AdresL1": ""}
        
        # 1. Szukamy NIPu
        nip_match = re.search(r'NIP:\s*(\d{10})', raw_text)
        if nip_match:
            result["NIP"] = nip_match.group(1)
            # Usuwamy linię z NIPem z dalszego procesowania
            lines = [l for l in lines if "NIP:" not in l]

        # 2. Ostatnia linia to zwykle Kod Pocztowy + Miasto
        if lines:
            result["AdresL1"] = lines[-1]
            # Spróbujmy dołączyć przedostatnią linię (ulica) jeśli to adres
            if len(lines) > 1:
                result["AdresL1"] = f"{lines[-2]}, {lines[-1]}"
                lines = lines[:-2]
            else:
                lines = lines[:-1]

        # 3. Reszta to Nazwa
        result["Nazwa"] = " ".join(lines)
        return result

    def _process_dynamic_rows(self, sheet, start_row_idx, start_col_idx):
        """Skanuje arkusz w dół od punktu startowego, rozpoznając zmiany nagłówków."""
        all_items = []
        current_row = start_row_idx
        
        # Inicjalny col_map (resetujemy przy każdym nowym nagłówku)
        col_map = {}

        while current_row <= sheet.max_row:
            row_cells = list(sheet.iter_rows(min_row=current_row, max_row=current_row))[0]
            # Bierzemy wartości z wiersza dla detekcji
            raw_vals = [str(c.value).strip().lower() if c.value is not None else "" for c in row_cells]
            val_0 = raw_vals[start_col_idx]
            
            # 1. Czy to wiersz z nagłówkami? (zawiera "nazwa usługi")
            if "nazwa usługi" in raw_vals or "nazwa uslugi" in raw_vals:
                col_map = {} # Reset mapowania
                try:
                    col_map["P_7"] = raw_vals.index("nazwa usługi")
                    # Szukamy różnych wariantów kolumn zależnie od układu
                    for ksef_tag, header_names in {
                        "P_8A": ["jm", "jednostka miary"],
                        "P_8B": ["ilość", "ilosc"],
                        "P_11": ["wartość sprzedaży netto w pln", "wartość sprzedaży netto pln", "wartosc netto pln"],
                        "P_12": ["stawka %", "stawka vat"],
                        "cena_eur": ["cena jednostkowa netto eur", "cena eur"],
                        "kwota_vat": ["kwota w pln", "kwota vat pln"]
                    }.items():
                        for name in header_names:
                            if name in raw_vals:
                                col_map[ksef_tag] = raw_vals.index(name)
                                break
                except Exception as e:
                    print(f"Błąd re-mappingu w wierszu {current_row}: {e}")
                
                current_row += 1
                continue

            # 2. Czy to wiersz z danymi (zaczyna się od liczby)?
            if val_0.replace('.', '').isdigit():
                if col_map:
                    item = {}
                    for ksef_field, col_idx in col_map.items():
                        item[ksef_field] = row_cells[col_idx].value
                    all_items.append(item)
            
            # 3. Warunek stopu
            if "razem" in val_0 or "suma" in val_0:
                break
                
            current_row += 1
        return all_items

    def create_xml_invoice(self, excel_path, mapping_name):
        mapping = self.load_mapping_config(mapping_name)
        rules = mapping['mapping']
        
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        sheet = wb.active 

        root = etree.Element("Faktura", nsmap=self.ns)
        
        # NAGŁÓWEK
        naglowek = etree.SubElement(root, "Naglowek")
        etree.SubElement(naglowek, "KodFormularza", kodSystemowy="FA (3)", wersjaSchemy="1-0E").text = "FA"
        etree.SubElement(naglowek, "WariantFormularza").text = "3"
        etree.SubElement(naglowek, "DataWytworzeniaFa").text = datetime.now().isoformat() + "Z"
        etree.SubElement(naglowek, "SystemInfo").text = "KSeF-Python-Client-V9"
        
        # SPRZEDAWCA
        p1 = etree.SubElement(root, "Podmiot1")
        dane_id1 = etree.SubElement(p1, "DaneIdentyfikacyjne")
        etree.SubElement(dane_id1, "NIP").text = self.cfg.nip
        etree.SubElement(dane_id1, "Nazwa").text = self.cfg.nazwa
        addr1 = etree.SubElement(p1, "Adres")
        etree.SubElement(addr1, "KodKraju").text = "PL"
        etree.SubElement(addr1, "AdresL1").text = self.cfg.adres
        
        # NABYWCA
        raw_buyer_text = str(self.get_value_from_excel(wb, sheet, rules['buyer']['FullBlock']) or "")
        buyer_data = self._parse_recipient_info(raw_buyer_text)
        
        p2 = etree.SubElement(root, "Podmiot2")
        dane_id2 = etree.SubElement(p2, "DaneIdentyfikacyjne")
        etree.SubElement(dane_id2, "NIP").text = buyer_data["NIP"]
        etree.SubElement(dane_id2, "Nazwa").text = buyer_data["Nazwa"]
        addr2 = etree.SubElement(p2, "Adres")
        etree.SubElement(addr2, "KodKraju").text = "PL"
        etree.SubElement(addr2, "AdresL1").text = buyer_data["AdresL1"]
        
        # SEKCJA FA
        fa = etree.SubElement(root, "Fa")
        etree.SubElement(fa, "KodWaluty").text = self.get_value_from_excel(wb, sheet, rules['header']['KodWaluty'])
        etree.SubElement(fa, "P_1").text = str(self.get_value_from_excel(wb, sheet, rules['header']['P_1']))
        etree.SubElement(fa, "P_2").text = str(self.get_value_from_excel(wb, sheet, rules['header']['P_2']))
        
        if 'P_6' in rules['header']:
            p6_val = self.get_value_from_excel(wb, sheet, rules['header']['P_6'])
            if p6_val: etree.SubElement(fa, "P_6").text = str(p6_val)

        # WIERSZE
        total_netto = 0
        total_vat = 0
        
        entry_anchor = wb.defined_names.get("entries_array")
        if entry_anchor:
            dest = list(entry_anchor.destinations)[0]
            start_sheet = wb[dest[0]]
            from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
            coord = dest[1].replace('$', '')
            start_row = int(re.search(r'\d+', coord).group())
            start_col = column_index_from_string(re.search(r'[A-Z]+', coord).group())
            
            # Skanujemy w górę odentries_array aby znaleźć nagłówki pierwszej tabeli
            items = self._process_dynamic_rows(start_sheet, start_row - 1, start_col - 1)
            
            for idx, item in enumerate(items, 1):
                wiersz = etree.SubElement(fa, "FaWiersz")
                etree.SubElement(wiersz, "NrWierszaFa").text = str(idx)
                etree.SubElement(wiersz, "P_7").text = str(item.get("P_7", ""))
                etree.SubElement(wiersz, "P_8A").text = str(item.get("P_8A", ""))
                etree.SubElement(wiersz, "P_8B").text = str(item.get("P_8B", ""))
                
                netto = float(item.get("P_11") or 0)
                vat = float(item.get("kwota_vat") or 0)
                total_netto += netto
                total_vat += vat
                
                etree.SubElement(wiersz, "P_11").text = f"{netto:.2f}"
                etree.SubElement(wiersz, "P_12").text = str(item.get("P_12", "23")).replace('%', '')
                
                if item.get("cena_eur"):
                    desc = etree.SubElement(fa, "DodatkowyOpis")
                    etree.SubElement(desc, "NrWiersza").text = str(idx)
                    etree.SubElement(desc, "Klucz").text = "Cena jednostkowa netto w EUR"
                    etree.SubElement(desc, "Wartosc").text = str(item["cena_eur"])

        # SUMY (P_13_1, P_14_1, P_15)
        etree.SubElement(fa, "P_13_1").text = f"{total_netto:.2f}"
        etree.SubElement(fa, "P_14_1").text = f"{total_vat:.2f}"
        etree.SubElement(fa, "P_15").text = f"{(total_netto + total_vat):.2f}"

        adnotacje = etree.SubElement(fa, "Adnotacje")
        for tag in ["P_16", "P_17", "P_18", "P_18A", "P_23"]:
            etree.SubElement(adnotacje, tag).text = "2"
        etree.SubElement(fa, "RodzajFaktury").text = "VAT"

        # Zapis
        output_path = excel_path.replace(".xlsx", ".xml")
        tree = etree.ElementTree(root)
        tree.write(output_path, encoding="utf-8", xml_declaration=True, pretty_print=True)
        return output_path

        # Adnotacje (wymagane pola P_16...P_18 na 2 = NIE)
        adnotacje = etree.SubElement(fa, "Adnotacje")
        for tag in ["P_16", "P_17", "P_18", "P_18A", "P_23"]:
            etree.SubElement(adnotacje, tag).text = "2"
        etree.SubElement(fa, "RodzajFaktury").text = "VAT"

        output_path = excel_path.replace(".xlsx", ".xml")
        tree = etree.ElementTree(root)
        tree.write(output_path, encoding="utf-8", xml_declaration=True, pretty_print=True)
        
        return output_path

if __name__ == "__main__":
    # Mały test "na sucho" wymagałby fizycznego pliku, 
    # więc na razie tylko definiujemy klasę.
    pass
