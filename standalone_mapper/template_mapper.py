import openpyxl
import re
import json
from lxml import etree
from datetime import datetime

class TemplateMapper:
    """
    Silnik mapujący oparty na szablonie i Nazwanych Zakresach.
    Zaprojektowany specjalnie dla faktur działu ACC.
    """
    def __init__(self, technical_rules_path):
        with open(technical_rules_path, 'r', encoding='utf-8') as f:
            self.rules = json.load(f)
        self.ns = {
            None: "http://crd.gov.pl/wzor/2025/06/25/13775/",
            "etd": "http://crd.gov.pl/xml/schematy/dziedzinowe/mf/2022/01/05/eD/DefinicjeTypy/",
            "xsi": "http://www.w3.org/2001/XMLSchema-instance"
        }

    def get_named_range_value(self, wb, name, sheet_name="Faktura"):
        """Pobiera wartość z nazwanego zakresu (globalnego lub lokalnego)."""
        # 1. Spróbuj lokalny w arkuszu
        sheet = wb[sheet_name]
        d_name = sheet.defined_names.get(name)
        
        # 2. Spróbuj globalny
        if not d_name:
            d_name = wb.defined_names.get(name)
            
        if not d_name:
            return None
            
        try:
            dest = list(d_name.destinations)[0]
            s_name, coord = dest
            coord = coord.replace('$', '')
            
            if ':' in coord:
                # To jest tablica/zakres - zwróć wiersze
                rows = []
                for row in wb[s_name][coord]:
                    rows.append([cell.value for cell in row])
                return rows
            else:
                # To jest pojedyncza komórka
                return wb[s_name][coord].value
        except:
            return None

    def get_multi_row_entries(self, wb, name, sheet_name="Faktura", max_scan=50):
        """Pobiera wiersze zaczynając od nazwanego zakresu, skanując w dół."""
        # Najpierw znajdź start
        sheet = wb[sheet_name]
        d_name = sheet.defined_names.get(name) or wb.defined_names.get(name)
        if not d_name: return []

        try:
            dest = list(d_name.destinations)[0]
            s_name, coord = dest
            coord = coord.replace('$', '')
            start_coord = coord.split(':')[0] if ':' in coord else coord
            
            # Wyciągnij kolumnę i wiersz startowy
            match = re.search(r'([A-Z]+)(\d+)', start_coord)
            if not match: return []
            start_col_letter = match.group(1)
            start_row = int(match.group(2))
            
            # Określ szerokość na podstawie nazwanego zakresu (jeśli to zakres)
            end_col_letter = start_col_letter
            if ':' in coord:
                end_match = re.search(r'([A-Z]+)\d+', coord.split(':')[1])
                if end_match: end_col_letter = end_match.group(1)
            
            rows = []
            # Skanuj w dół
            for r in range(start_row, start_row + max_scan):
                # Wyciągnij kolumny od start do end
                import openpyxl.utils
                min_col = openpyxl.utils.column_index_from_string(start_col_letter)
                max_col = openpyxl.utils.column_index_from_string(end_col_letter)
                
                row_vals = []
                for c in range(min_col, max_col + 1):
                    col_letter = openpyxl.utils.get_column_letter(c)
                    row_vals.append(wb[s_name][f"{col_letter}{r}"].value)
                
                # Warunek stopu: "Razem", "Ogółem", "Metoda płatności" itp.
                # Nie zatrzymuj się, jeśli PIERWSZA kolumna jest liczbą (to znaczy, że to kolejna pozycja)
                first_val_str = str(row_vals[0]).strip() if row_vals[0] else ""
                
                check_str = (str(row_vals[0] or "") + " " + str(row_vals[1] or "")).lower()
                stop_words = ["razem", "ogółem", "total", "metoda płatności", "payment method", "konto bankowe"]
                
                # Jeśli pierwsza kolumna to liczba, ignoruj stop words (bo mogą być w opisie)
                is_line_item = False
                if first_val_str.replace('.', '').isdigit():
                    is_line_item = True
                
                if not is_line_item and any(word in check_str for word in stop_words):
                    break
                
                # Warunek stopu: Kolejny Lp jest mniejszy niż poprzedni lub nie jest liczbą (ale uważaj na wiersze opisu)
                # Na razie polegamy na "Razem" i filtracji w mapowaniu.
                    
                rows.append(row_vals)
            return rows
        except Exception as e:
            print(f"Błąd get_multi_row_entries: {e}")
            return []

    def parse_recipient(self, raw_text):
        """Rozbija blok tekstu nabywcy na części z uwzględnieniem wielowierszowej nazwy."""
        if not raw_text: return {"Nazwa": "", "NIP": "", "Adres": "", "KodKraju": "PL"}
        lines = [l.strip() for l in raw_text.split('\n') if l.strip()]
        
        result = {"Nazwa": "", "NIP": "", "Adres": "", "KodKraju": "PL"}
        
        # Ekstrakcja NIP / Tax ID
        # Szukamy: NIP: PL123... lub Tax ID: FR 123...
        # Pattern łapie: (NIP|Tax ID)[:\s]*([A-Z]{2})?\s*([0-9A-Z\s]+) 
        # Ale uprośćmy: szukamy ciągu cyframi/literami po słowach kluczowych
        
        # 1. NIP polski
        nip_match = re.search(r'NIP:\s*(?:PL)?\s*(\d{9,11})', raw_text, re.IGNORECASE)
        
        # 2. Tax ID zagraniczny (np. Tax ID: FR 93319378832)
        tax_id_match = re.search(r'Tax ID:\s*([A-Z]{2})\s*([0-9A-Z]+)', raw_text, re.IGNORECASE)
        
        if nip_match:
            result["NIP"] = nip_match.group(1)
            result["KodKraju"] = "PL"
            lines = [l for l in lines if "NIP:" not in l.upper()]
        elif tax_id_match:
            result["KodKraju"] = tax_id_match.group(1)
            result["NIP"] = tax_id_match.group(2) # Zwracamy sam numer bez kodu kraju jeśli KSeF tak woli, LUB całość?
            # KSeF w polu NIP dla zagranicznych podmiotów oczekuje samego numeru, kod kraju idzie osobno.
            # Ale w elemencie NIP schemy FA(3) jest pattern... 
            # Dla Podmiot2 (Nabywca) -> DaneIdentyfikacyjne -> NIP (do 30 znaków).
            # Jeśli to podmiot zagraniczny, to KodKraju jest w Adres -> KodKraju.
            # Więc do pola NIP wpisujemy numer identyfikacyjny.
            
            # Usuń linię z Tax ID
            lines = [l for l in lines if "Tax ID:" not in l] # Case sensitive? better check manually
            # Bardziej precyzyjne usuwanie:
            new_lines = []
            for l in lines:
                if "Tax ID:" in l:
                    # Może być w środku linii adresu, np. "Adres..., Tax ID: ..."
                    # Wtedy trzeba wyciąć tylko fragment, ale to trudne.
                    # W podanym przykładzie: "..., France, Tax ID: FR..., -"
                    # Załóżmy że jeśli znajdziemy w linii, to NIP już mamy, a reszta to adres.
                    # Ale lines pochodzi ze splita po \n.
                    # Oryginalny tekst: AdresL1 w Excelu może mieć \n lub nie.
                    pass
                new_lines.append(l)
            # Przywróćmy lines, bo logika usuwania z treści adresu jest ryzykowna.
            # Jeśli "Tax ID" jest w tej samej linii co adres, to parse_recipient (który dzieli po \n)
            # może mieć to w jednej linii.
            
        # Jeśli Tax ID wyciągnęliśmy regexem z raw_text, to musimy uważać żeby nie został w "Adresie" jako śmieć.
        # Spróbujmy wyczyścić raw_text z NIP/Tax ID przed dzieleniem na linie dla adresu/nazwy?
        # Nie, bo Nazwa jest na początku.
        
        # Prostsze podejście: Najpierw wyciągnij dane, potem zgaduj nazwę/adres.
        
        if len(lines) >= 2:
            # Sprawdź czy druga linia to ciąg dalszy nazwy
            if not re.search(r'\d{2}-\d{3}|ul\.|al\.|os\.|pl\.|rue|avenue|str\.', lines[1], re.IGNORECASE):
                result["Nazwa"] = f"{lines[0]}\n{lines[1]}"
                result["Adres"] = ", ".join(lines[2:])
            else:
                result["Nazwa"] = lines[0]
                result["Adres"] = ", ".join(lines[1:])
        elif len(lines) == 1:
            result["Nazwa"] = lines[0]
            
        # Clean address from Tax ID if present
        if result["Adres"] and "Tax ID:" in result["Adres"]:
             result["Adres"] = re.sub(r'Tax ID:\s*[A-Z]{2}\s*[0-9A-Z]+', '', result["Adres"]).strip().strip(',').strip()

        return result

    def format_date(self, val):
        if isinstance(val, datetime):
            return val.strftime("%Y-%m-%d")
        if not val: return ""
        # Próba wyciągnięcia daty ze stringa YYYY-MM-DD
        match = re.search(r'(\d{4}-\d{2}-\d{2})', str(val))
        return match.group(1) if match else str(val)

    def extract_invoice_no(self, title):
        """Wyciąga numer faktury z np. 'Faktura VAT/invoice 2025-10-FV03'"""
        if not title: return ""
        # Szukamy wzorca rrrr-mm-FVxx
        match = re.search(r'(\d{4}-\d{2}-FV\d+)', title, re.IGNORECASE)
        return match.group(1) if match else title

    def extract_data(self, excel_path, mapping_type="ACC_PLN_PROSTA"):
        """
        Parses Excel and returns structured data for UI.
        Does NOT generate XML.
        """
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        
        # Default empty result
        result = {
            "status": "NEW",
            "file": excel_path,
            "template": mapping_type,
            "p1": {}, "p2": {}, "inv": {}, "payment": "", "footer": ""
        }
        
        # P1 is static in rules
        p1_cfg = self.rules.get('podmiot1_static', {})
        result['p1'] = {
            "nip": p1_cfg.get('NIP', ''),
            "country": p1_cfg.get('KodKraju', 'PL'),
            "name": p1_cfg.get('Nazwa', ''),
            "addr": p1_cfg.get('AdresL1', '')
        }
        
        # Common data extraction logic (similar to _map methods)
        data = {}
        buyer = {}
        totals = {"net": 0.0, "vat": 0.0, "gross": 0.0, "curr": "PLN"}
        
        try:
            # Shared named ranges
            raw_title = self.get_named_range_value(wb, 'note_title')
            raw_date_issue = self.get_named_range_value(wb, 'date_issue')
            raw_date_sale = self.get_named_range_value(wb, 'date_sale')
            raw_recipient = self.get_named_range_value(wb, 'recipient_info')
            raw_payment = self.get_named_range_value(wb, 'payment_method_info')
            
            # Invoice Basic Info
            result['inv']['no'] = self.extract_invoice_no(raw_title)
            result['inv']['date_issued'] = self.format_date(raw_date_issue)
            result['inv']['date_service'] = self.format_date(raw_date_sale)
            result['payment'] = str(raw_payment) if raw_payment else ""
            result['footer'] = p1_cfg.get('StopkaFaktury', '')
            
            # Buyer Parsing
            buyer = self.parse_recipient(raw_recipient)
            result['p2'] = {
                "nip": buyer.get('NIP', ''),
                "country": buyer.get('KodKraju', 'PL'),
                "name": buyer.get('Nazwa', '').replace('\n', ' '),
                "jst": "2", "gv": "2", # Defaults
                "addr": buyer.get('Adres', '') # Adding Address for UI
            }
            
            # Totals Calculation based on mapping type
            if mapping_type == "ACC_PLN_PROSTA":
                entries = self.get_named_range_value(wb, 'entries_array')
                if entries:
                    for row in entries:
                        if not row[0]: continue
                        try:
                            netto = float(row[4] or 0)
                            vat_val = float(row[6] or 0)
                            totals['net'] += netto
                            totals['vat'] += vat_val
                        except: pass
                totals['curr'] = "PLN"
                
            elif mapping_type in ["ACC_MULTI", "ACC_EUR"]:
                entries = self.get_multi_row_entries(wb, 'entries_array')
                # Determine currency
                is_eur = (mapping_type == "ACC_EUR")
                totals['curr'] = "EUR" if is_eur else "PLN" # MULTI is PLN report
                
                if entries:
                    for row in entries:
                        if not row or all(v is None for v in row): continue
                         # Skip headers logic (simplified)
                        if "jednost." in str(row[2]).lower(): continue
                        
                        first_col = str(row[0]).replace('.', '')
                        if first_col.isdigit():
                            try:
                                # Logic depends on column index which varies slightly or is same?
                                # PROSTA: 4=Netto, 6=VAT
                                # MULTI: 4=Cena EUR, 5=Netto EUR, 6=Netto PLN, 8=Vat PLN
                                # EUR: 5=EUR Unit, 6=EUR Netto
                                
                                if mapping_type == "ACC_MULTI":
                                    netto = float(row[6] or 0) # PLN Netto
                                    vat_val = float(row[8] or 0) # PLN Vat
                                    totals['net'] += netto
                                    totals['vat'] += vat_val
                                    
                                elif mapping_type == "ACC_EUR":
                                    netto = float(row[6] or 0) # EUR Netto
                                    rate_str = str(row[7]).replace('%', '').strip()
                                    vat_rate = float(rate_str) if rate_str and rate_str.replace('.','').isdigit() else 0
                                    vat_val = netto * (vat_rate / 100.0)
                                    totals['net'] += netto
                                    totals['vat'] += vat_val
                            except: pass

            totals['gross'] = totals['net'] + totals['vat']
            
            # Assign totals formatted
            result['inv']['curr'] = totals['curr']
            result['inv']['net'] = f"{totals['net']:.2f}"
            result['inv']['vat'] = f"{totals['vat']:.2f}"
            result['inv']['gross'] = f"{totals['gross']:.2f}"
            
            result['status'] = "ZAŁADOWANY" # Success state
            
        except Exception as e:
            result['status'] = "BŁĄD"
            print(f"Extraction error: {e}")
            
        return result

    def create_xml(self, excel_path, output_path, mapping_type="ACC_PLN_PROSTA"):
        """Główny dispatcher wybierający odpowiednią logikę mapowania."""
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        
        if mapping_type == "ACC_PLN_PROSTA":
            return self._map_acc_pln_prosta(wb, output_path)
        elif mapping_type == "ACC_MULTI":
            return self._map_acc_multi(wb, output_path)
        elif mapping_type == "ACC_EUR":
            return self._map_acc_eur(wb, output_path)
        else:
            raise ValueError(f"Nieznany typ mapowania: {mapping_type}")

    def _map_acc_pln_prosta(self, wb, output_path):
        """Opracowana wcześniej logika dla faktur PLN Prosta."""
        data = {
            "invoice_title": self.get_named_range_value(wb, 'note_title'),
            "date_issue": self.get_named_range_value(wb, 'date_issue'),
            "date_sale": self.get_named_range_value(wb, 'date_sale'),
            "raw_recipient": self.get_named_range_value(wb, 'recipient_info'),
            "entries": self.get_named_range_value(wb, 'entries_array'),
            "bank_account": self.get_named_range_value(wb, 'beneficiary_account'),
            "payment_method": self.get_named_range_value(wb, 'payment_method_info')
        }

        invoice_no = self.extract_invoice_no(data['invoice_title'])
        buyer = self.parse_recipient(data['raw_recipient'])
        
        # 2. Budowa XML
        root = etree.Element("Faktura", nsmap=self.ns)
        
        # Naglowek
        nag = etree.SubElement(root, "Naglowek")
        defaults = self.rules['system_defaults']
        etree.SubElement(nag, "KodFormularza", kodSystemowy=defaults['kodSystemowy'], wersjaSchemy=defaults['wersjaSchemy']).text = defaults['KodFormularza']
        etree.SubElement(nag, "WariantFormularza").text = str(defaults['WariantFormularza'])
        etree.SubElement(nag, "DataWytworzeniaFa").text = datetime.now().isoformat() + "Z"
        etree.SubElement(nag, "SystemInfo").text = "KSeF-ACC-Mapper-V2"

        # Podmiot1 (Sprzedawca - BNP)
        p1_cfg = self.rules['podmiot1_static']
        p1 = etree.SubElement(root, "Podmiot1")
        etree.SubElement(p1, "PrefiksPodatnika").text = p1_cfg['PrefiksPodatnika']
        d_id1 = etree.SubElement(p1, "DaneIdentyfikacyjne")
        etree.SubElement(d_id1, "NIP").text = p1_cfg['NIP']
        etree.SubElement(d_id1, "Nazwa").text = p1_cfg['Nazwa']
        addr1 = etree.SubElement(p1, "Adres")
        etree.SubElement(addr1, "KodKraju").text = p1_cfg['KodKraju']
        etree.SubElement(addr1, "AdresL1").text = p1_cfg['AdresL1']

        # Podmiot2 (Nabywca)
        p2 = etree.SubElement(root, "Podmiot2")
        d_id2 = etree.SubElement(p2, "DaneIdentyfikacyjne")
        etree.SubElement(d_id2, "NIP").text = buyer['NIP']
        etree.SubElement(d_id2, "Nazwa").text = buyer['Nazwa']
        addr2 = etree.SubElement(p2, "Adres")
        etree.SubElement(addr2, "KodKraju").text = "PL"
        etree.SubElement(addr2, "AdresL1").text = buyer['Adres']
        etree.SubElement(p2, "JST").text = "2"
        etree.SubElement(p2, "GV").text = "2"

        # Fa
        fa = etree.SubElement(root, "Fa")
        etree.SubElement(fa, "KodWaluty").text = "PLN"
        etree.SubElement(fa, "P_1").text = self.format_date(data['date_issue'])
        etree.SubElement(fa, "P_2").text = invoice_no
        etree.SubElement(fa, "P_6").text = self.format_date(data['date_sale'])
        
        # 1. Przetwórz dane (obliczenia) ZANIM dodasz tagi do XML
        is_exempt = False
        total_netto = 0
        total_vat = 0
        processed_items = []

        for row in data['entries']:
            if not row[0]: continue
            netto = float(row[4] or 0)
            total_netto += netto
            rate = str(row[5]).lower()
            vat = 0
            if 'zw' in rate:
                is_exempt = True
            else:
                vat = float(row[6] or 0)
                total_vat += vat
            
            processed_items.append({
                "P_7": str(row[1]),
                "P_8A": str(row[2]),
                "P_8B": str(row[3]),
                "P_9A": f"{netto}",
                "P_11": f"{netto}",
                "P_12": "zw" if is_exempt and 'zw' in rate else rate.replace('%', '').strip(),
                "P_11Vat": f"{vat}" if vat > 0 else None
            })

        # 2. DODAJ TAGI W ŚCISŁEJ KOLEJNOŚCI (Schema FA 3)
        
        # Sumy (P_13, P_14...)
        if is_exempt:
            etree.SubElement(fa, "P_13_7").text = f"{total_netto}"
        else:
            etree.SubElement(fa, "P_13_1").text = f"{total_netto}"
            etree.SubElement(fa, "P_14_1").text = f"{total_vat}"
            
        etree.SubElement(fa, "P_15").text = f"{(total_netto + total_vat):.2f}"

        # Adnotacje
        adnot = etree.SubElement(fa, "Adnotacje")
        for tag in ["P_16", "P_17", "P_18", "P_18A"]:
            etree.SubElement(adnot, tag).text = "2"
            
        if is_exempt:
            zw = etree.SubElement(adnot, "Zwolnienie")
            etree.SubElement(zw, "P_19").text = "1"
            etree.SubElement(zw, "P_19A").text = self.rules['vat_rules']['default_exempt_basis']
        else:
            zw = etree.SubElement(adnot, "Zwolnienie")
            etree.SubElement(zw, "P_19N").text = "1"

        nst = etree.SubElement(adnot, "NoweSrodkiTransportu")
        etree.SubElement(nst, "P_22N").text = "1"
        etree.SubElement(adnot, "P_23").text = "2"
        pm = etree.SubElement(adnot, "PMarzy")
        etree.SubElement(pm, "P_PMarzyN").text = "1"

        etree.SubElement(fa, "RodzajFaktury").text = "VAT"

        # Wiersze (FaWiersz) - Muszą być PO Adnotacjach i RodzajFaktury
        for idx, item in enumerate(processed_items, 1):
            w = etree.SubElement(fa, "FaWiersz")
            etree.SubElement(w, "NrWierszaFa").text = str(idx)
            etree.SubElement(w, "P_7").text = item["P_7"]
            etree.SubElement(w, "P_8A").text = item["P_8A"]
            etree.SubElement(w, "P_8B").text = item["P_8B"]
            etree.SubElement(w, "P_9A").text = item["P_9A"]
            etree.SubElement(w, "P_11").text = item["P_11"]
            if item["P_11Vat"]:
                etree.SubElement(w, "P_11Vat").text = item["P_11Vat"]
            etree.SubElement(w, "P_12").text = item["P_12"]

        # Platnosc
        plat = etree.SubElement(fa, "Platnosc")
        term = etree.SubElement(plat, "TerminPlatnosci")
        desc = etree.SubElement(term, "TerminOpis")
        etree.SubElement(desc, "Ilosc").text = "10"
        etree.SubElement(desc, "Jednostka").text = "dni/days"
        etree.SubElement(desc, "ZdarzeniePoczatkowe").text = str(data['payment_method'])
        
        etree.SubElement(plat, "FormaPlatnosci").text = "6"
        rb = etree.SubElement(plat, "RachunekBankowy")
        etree.SubElement(rb, "NrRB").text = str(data['bank_account']).replace(' ', '')
        etree.SubElement(rb, "NazwaBanku").text = p1_cfg['Nazwa']

        # Stopka
        stopka = etree.SubElement(root, "Stopka")
        inf = etree.SubElement(stopka, "Informacje")
        etree.SubElement(inf, "StopkaFaktury").text = p1_cfg.get('StopkaFaktury', '')
        rej = etree.SubElement(stopka, "Rejestry")
        etree.SubElement(rej, "PelnaNazwa").text = p1_cfg['Nazwa']
        etree.SubElement(rej, "KRS").text = "0000245000"

        # Zapisz
        tree = etree.ElementTree(root)
        tree.write(output_path, encoding="utf-8", xml_declaration=True, pretty_print=True)
        return output_path
    def _map_acc_multi(self, wb, output_path):
        """Opracowana logika dla faktur wielowalutowych (ACC_MULTI)."""
        data = {
            "invoice_title": self.get_named_range_value(wb, 'note_title'),
            "date_issue": self.get_named_range_value(wb, 'date_issue'),
            "date_sale": self.get_named_range_value(wb, 'date_sale'),
            "raw_recipient": self.get_named_range_value(wb, 'recipient_info'),
            "entries": self.get_multi_row_entries(wb, 'entries_array'), # Skanowanie w dół
            "bank_account": self.get_named_range_value(wb, 'beneficiary_account'),
            "payment_method": self.get_named_range_value(wb, 'payment_method_info')
        }

        # Ekstrakcja kursu waluty (z palca z C34 lub szukanie nad tabelą)
        exchange_rate = 0.0
        sheet = wb["Faktura"]
        val_c34 = sheet["C34"].value
        try:
            exchange_rate = float(val_c34) if val_c34 else 0.0
        except:
            # Skan kolumny C powyżej entries_array (zakładamy start w wierszu 18)
            for r in range(12, 18):
                v = sheet[f"C{r}"].value
                try:
                    if v and float(v) > 0.1: # Prawdopodobny kurs
                        exchange_rate = float(v)
                        break
                except: continue

        invoice_no = self.extract_invoice_no(data['invoice_title'])
        buyer = self.parse_recipient(data['raw_recipient'])
        
        # 2. Budowa XML
        root = etree.Element("Faktura", nsmap=self.ns)
        
        # Naglowek
        nag = etree.SubElement(root, "Naglowek")
        defaults = self.rules['system_defaults']
        etree.SubElement(nag, "KodFormularza", kodSystemowy=defaults['kodSystemowy'], wersjaSchemy=defaults['wersjaSchemy']).text = defaults['KodFormularza']
        etree.SubElement(nag, "WariantFormularza").text = str(defaults['WariantFormularza'])
        etree.SubElement(nag, "DataWytworzeniaFa").text = datetime.now().isoformat() + "Z"
        etree.SubElement(nag, "SystemInfo").text = "KSeF-ACC-Mapper-Multi-V2"

        # Podmiot1 (Sprzedawca - BNP)
        p1_cfg = self.rules['podmiot1_static']
        p1 = etree.SubElement(root, "Podmiot1")
        etree.SubElement(p1, "PrefiksPodatnika").text = p1_cfg['PrefiksPodatnika']
        d_id1 = etree.SubElement(p1, "DaneIdentyfikacyjne")
        etree.SubElement(d_id1, "NIP").text = p1_cfg['NIP']
        etree.SubElement(d_id1, "Nazwa").text = p1_cfg['Nazwa']
        addr1 = etree.SubElement(p1, "Adres")
        etree.SubElement(addr1, "KodKraju").text = p1_cfg['KodKraju']
        etree.SubElement(addr1, "AdresL1").text = p1_cfg['AdresL1']

        # Podmiot2 (Nabywca)
        p2 = etree.SubElement(root, "Podmiot2")
        d_id2 = etree.SubElement(p2, "DaneIdentyfikacyjne")
        etree.SubElement(d_id2, "NIP").text = buyer['NIP']
        etree.SubElement(d_id2, "Nazwa").text = buyer['Nazwa']
        addr2 = etree.SubElement(p2, "Adres")
        etree.SubElement(addr2, "KodKraju").text = "PL"
        etree.SubElement(addr2, "AdresL1").text = buyer['Adres']
        etree.SubElement(p2, "JST").text = "2"
        etree.SubElement(p2, "GV").text = "2"

        # Fa
        fa = etree.SubElement(root, "Fa")
        etree.SubElement(fa, "KodWaluty").text = "PLN" # KSeF Raportujemy w PLN mimo kursu
        etree.SubElement(fa, "P_1").text = self.format_date(data['date_issue'])
        etree.SubElement(fa, "P_2").text = invoice_no
        etree.SubElement(fa, "P_6").text = self.format_date(data['date_sale'])
        
        # Przetwarzanie danych
        is_exempt = False
        total_netto = 0.0
        total_vat = 0.0
        processed_items = []

        # Kolumny ACC_MULTI:
        # 0: Lp, 1: Nazwa, 2: Jm, 3: Ilość, 4: Cena EUR, 5: Netto EUR, 6: Netto PLN, 7: VAT %, 8: Kwota VAT PLN
        
        current_item = None
        for row in data['entries']:
            if not row or all(v is None for v in row): continue
            
            # Czy to wiersz z nagłówkiem tabeli?
            if any(word in str(row[2]).lower() for word in ["jednost.", "miary"]) or \
               any(word in str(row[1]).lower() for word in ["nazwa usługi"]):
                continue

            first_col = str(row[0]).strip().replace('.', '')
            
            # PRZYPADEK 1: Nowa pozycja (ma Lp)
            if first_col.isdigit():
                # Jeśli mieliśmy poprzedni niedokończony item, zapisz go
                if current_item:
                    processed_items.append(current_item)
                
                try:
                    netto_pln = float(row[6] or 0)
                    total_netto += netto_pln
                    rate = str(row[7]).lower()
                    vat_pln = 0.0
                    if 'zw' in rate: is_exempt = True
                    else:
                        vat_pln = float(row[8] or 0)
                        total_vat += vat_pln
                    
                    qty = float(row[3] or 1)
                    unit_price_pln = netto_pln / qty if qty != 0 else netto_pln
                    
                    current_item = {
                        "P_7": str(row[1]) if row[1] else "",
                        "P_8A": str(row[2]) if row[2] else "",
                        "P_8B": str(int(qty)) if qty == int(qty) else f"{qty}",
                        "P_9A": f"{unit_price_pln:.2f}",
                        "P_11": f"{netto_pln:.2f}",
                        "P_11Vat": f"{vat_pln:.2f}" if vat_pln > 0 else None,
                        "P_12": "zw" if 'zw' in rate else rate.replace('%', '').strip(),
                        "EUR_Unit": float(row[4] or 0) if row[4] and row[5] else None,
                        "has_eur": row[5] is not None
                    }
                except:
                    current_item = None
                    continue
            
            # PRZYPADEK 2: Wiersz bez Lp, ale z opisem (często pod wierszem z danymi)
            elif row[1] and current_item and not current_item["P_7"]:
                current_item["P_7"] = str(row[1])
            elif row[1] and current_item:
                # Kontynuacja opisu
                current_item["P_7"] += "\n" + str(row[1])

        # Zapisz ostatni item
        if current_item:
            processed_items.append(current_item)

        # Sumy
        if is_exempt:
            etree.SubElement(fa, "P_13_7").text = f"{total_netto:.2f}"
        else:
            etree.SubElement(fa, "P_13_1").text = f"{total_netto:.2f}"
            etree.SubElement(fa, "P_14_1").text = f"{total_vat:.2f}"
            
        etree.SubElement(fa, "P_15").text = f"{(total_netto + total_vat):.2f}"

        # Adnotacje
        adnot = etree.SubElement(fa, "Adnotacje")
        for tag in ["P_16", "P_17", "P_18", "P_18A"]:
            etree.SubElement(adnot, tag).text = "2"
            
        if is_exempt:
            zw = etree.SubElement(adnot, "Zwolnienie")
            etree.SubElement(zw, "P_19").text = "1"
            etree.SubElement(zw, "P_19A").text = self.rules['vat_rules']['default_exempt_basis']
        else:
            zw = etree.SubElement(adnot, "Zwolnienie")
            etree.SubElement(zw, "P_19N").text = "1"

        nst = etree.SubElement(adnot, "NoweSrodkiTransportu")
        etree.SubElement(nst, "P_22N").text = "1"
        etree.SubElement(adnot, "P_23").text = "2"
        pm = etree.SubElement(adnot, "PMarzy")
        etree.SubElement(pm, "P_PMarzyN").text = "1"

        etree.SubElement(fa, "RodzajFaktury").text = "VAT"

        # DodatkowyOpis (Multi-currency info)
        # 1. Ceny jednostkowe EUR (tylko dla tych co mają has_eur)
        idx_desc = 1
        for item in processed_items:
            if item.get("has_eur") and item["EUR_Unit"]:
                dop = etree.SubElement(fa, "DodatkowyOpis")
                etree.SubElement(dop, "NrWiersza").text = str(idx_desc)
                etree.SubElement(dop, "Klucz").text = "Cena jednostkowa netto w EUR"
                etree.SubElement(dop, "Wartosc").text = f"{item['EUR_Unit']:.2f}".replace('.', ',')
                idx_desc += 1
        
        # 2. Kurs PLN/EUR
        if exchange_rate > 0:
            dop = etree.SubElement(fa, "DodatkowyOpis")
            etree.SubElement(dop, "NrWiersza").text = "1" # Reset wiersza wg wzorca dla klucza Kurs
            etree.SubElement(dop, "Klucz").text = "Kurs PLN/EUR"
            etree.SubElement(dop, "Wartosc").text = f"{exchange_rate:.6f}".replace('.', ',')

        # Wiersze (FaWiersz)
        for idx, item in enumerate(processed_items, 1):
            w = etree.SubElement(fa, "FaWiersz")
            etree.SubElement(w, "NrWierszaFa").text = str(idx)
            etree.SubElement(w, "P_7").text = item["P_7"]
            etree.SubElement(w, "P_8A").text = item["P_8A"]
            etree.SubElement(w, "P_8B").text = item["P_8B"]
            etree.SubElement(w, "P_9A").text = item["P_9A"]
            etree.SubElement(w, "P_11").text = item["P_11"]
            if item["P_11Vat"]:
                etree.SubElement(w, "P_11Vat").text = item["P_11Vat"]
            etree.SubElement(w, "P_12").text = item["P_12"]
            # KursWaluty dodajemy tylko gdy NIE MA EUR w DodatkowyOpis dla tego wiersza (zgodnie z wzorcem KPMG)
            if exchange_rate > 0 and not item.get("has_eur"):
                etree.SubElement(w, "KursWaluty").text = f"{exchange_rate:.4f}"

        # Platnosc
        plat = etree.SubElement(fa, "Platnosc")
        term = etree.SubElement(plat, "TerminPlatnosci")
        desc = etree.SubElement(term, "TerminOpis")
        etree.SubElement(desc, "Ilosc").text = "10"
        etree.SubElement(desc, "Jednostka").text = "dni/days"
        etree.SubElement(desc, "ZdarzeniePoczatkowe").text = str(data['payment_method'])
        
        etree.SubElement(plat, "FormaPlatnosci").text = "6"
        rb = etree.SubElement(plat, "RachunekBankowy")
        etree.SubElement(rb, "NrRB").text = str(data['bank_account']).replace(' ', '')
        etree.SubElement(rb, "NazwaBanku").text = p1_cfg['Nazwa']

        # Stopka
        stopka = etree.SubElement(root, "Stopka")
        inf = etree.SubElement(stopka, "Informacje")
        etree.SubElement(inf, "StopkaFaktury").text = p1_cfg.get('StopkaFaktury', '')
        rej = etree.SubElement(stopka, "Rejestry")
        etree.SubElement(rej, "PelnaNazwa").text = p1_cfg['Nazwa']
        etree.SubElement(rej, "KRS").text = "0000245000"

        # Zapisz
        tree = etree.ElementTree(root)
        tree.write(output_path, encoding="utf-8", xml_declaration=True, pretty_print=True)
        return output_path

    def _map_acc_eur(self, wb, output_path):
        """Logika dla faktur w walucie EUR (ACC_EUR)."""
        data = {
            "invoice_title": self.get_named_range_value(wb, 'note_title'),
            "date_issue": self.get_named_range_value(wb, 'date_issue'),
            "date_sale": self.get_named_range_value(wb, 'date_sale'),
            "raw_recipient": self.get_named_range_value(wb, 'recipient_info'),
            "entries": self.get_multi_row_entries(wb, 'entries_array'),
            "bank_account": self.get_named_range_value(wb, 'beneficiary_account'),
            "payment_method": self.get_named_range_value(wb, 'payment_method_info')
        }

        # Pobierz kurs z nazwanego zakresu exchange_rate
        exchange_rate = 0.0
        try:
            val_rate = self.get_named_range_value(wb, 'exchange_rate')
            if val_rate:
                exchange_rate = float(val_rate)
        except: pass

        invoice_no = self.extract_invoice_no(data['invoice_title'])
        buyer = self.parse_recipient(data['raw_recipient'])
        
        # XML
        root = etree.Element("Faktura", nsmap=self.ns)
        
        # Naglowek
        nag = etree.SubElement(root, "Naglowek")
        defaults = self.rules['system_defaults']
        etree.SubElement(nag, "KodFormularza", kodSystemowy=defaults['kodSystemowy'], wersjaSchemy=defaults['wersjaSchemy']).text = defaults['KodFormularza']
        etree.SubElement(nag, "WariantFormularza").text = str(defaults['WariantFormularza'])
        etree.SubElement(nag, "DataWytworzeniaFa").text = datetime.now().isoformat() + "Z"
        etree.SubElement(nag, "SystemInfo").text = "KSeF-ACC-Mapper-EUR-V1"

        # Podmiot1 (BNP)
        p1_cfg = self.rules['podmiot1_static']
        p1 = etree.SubElement(root, "Podmiot1")
        etree.SubElement(p1, "PrefiksPodatnika").text = p1_cfg['PrefiksPodatnika']
        d_id1 = etree.SubElement(p1, "DaneIdentyfikacyjne")
        etree.SubElement(d_id1, "NIP").text = p1_cfg['NIP']
        etree.SubElement(d_id1, "Nazwa").text = p1_cfg['Nazwa']
        addr1 = etree.SubElement(p1, "Adres")
        etree.SubElement(addr1, "KodKraju").text = p1_cfg['KodKraju']
        etree.SubElement(addr1, "AdresL1").text = p1_cfg['AdresL1']

        # Podmiot2 (Nabywca)
        p2 = etree.SubElement(root, "Podmiot2")
        d_id2 = etree.SubElement(p2, "DaneIdentyfikacyjne")
        etree.SubElement(d_id2, "NIP").text = buyer['NIP']
        etree.SubElement(d_id2, "Nazwa").text = buyer['Nazwa']
        addr2 = etree.SubElement(p2, "Adres")
        etree.SubElement(addr2, "KodKraju").text = buyer.get("KodKraju", "PL")
        etree.SubElement(addr2, "AdresL1").text = buyer['Adres']
        # Domyślne wartości JST/GV
        etree.SubElement(p2, "JST").text = "2"
        etree.SubElement(p2, "GV").text = "2"

        # Fa
        fa = etree.SubElement(root, "Fa")
        etree.SubElement(fa, "KodWaluty").text = "EUR"
        etree.SubElement(fa, "P_1").text = self.format_date(data['date_issue'])
        etree.SubElement(fa, "P_2").text = invoice_no
        etree.SubElement(fa, "P_6").text = self.format_date(data['date_sale'])
        
        # Przetwarzanie pozycji
        total_netto_eur = 0.0
        total_vat_eur = 0.0
        is_exempt = False
        is_np = False
        processed_items = []

        current_item = None
        for row in data['entries']:
            if not row or all(v is None for v in row): continue
            
            # Pomiń nagłówki
            if any(word in str(row[2]).lower() for word in ["jednost.", "miary"]) or \
               any(word in str(row[1]).lower() for word in ["nazwa usługi"]):
                continue

            first_col = str(row[0]).strip().replace('.', '')
            
            if first_col.isdigit():
                if current_item: processed_items.append(current_item)
                
                try:
                    # Col 4 = PLN Netto, Col 5 = EUR Unit, Col 6 = EUR Netto
                    pln_netto = float(row[4] or 0)
                    eur_unit = float(row[5] or 0)
                    eur_netto = float(row[6] or 0)
                    qty = float(row[3] or 1)
                    
                    rate_str = str(row[7]).lower()
                    
                    vat_eur_val = 0.0
                    p_12 = ""
                    
                    if "poza vat" in rate_str or "outside vat" in rate_str:
                        p_12 = "np"
                        is_np = True
                    elif "zw" in rate_str:
                        p_12 = "zw"
                        is_exempt = True
                    else:
                        # Standard VAT (np. 23%)
                        rate_clean = rate_str.replace('%', '').strip()
                        p_12 = rate_clean
                        try:
                            rate_val = float(rate_clean)
                            if rate_val > 0:
                                vat_eur_val = eur_netto * (rate_val / 100.0)
                        except: pass
                    
                    total_netto_eur += eur_netto
                    total_vat_eur += vat_eur_val
                    
                    current_item = {
                        "P_7": str(row[1]) if row[1] else "",
                        "P_8A": str(row[2]) if row[2] else "",
                        "P_8B": str(int(qty)) if qty == int(qty) else f"{qty}",
                        "P_9A": f"{eur_unit:.2f}",
                        "P_11": f"{eur_netto:.2f}",
                        "P_11Vat": f"{vat_eur_val:.2f}" if vat_eur_val > 0 else None,
                        "P_12": p_12,
                        "PLN_Netto": pln_netto
                    }
                except: 
                    current_item = None
                    continue
            
            elif row[1] and current_item and not current_item["P_7"]:
                current_item["P_7"] = str(row[1])
            elif row[1] and current_item:
                current_item["P_7"] += "\n" + str(row[1])

        if current_item: processed_items.append(current_item)

        # KursWalutyZ (wymagane przy walucie obcej jesli VAT rozliczany w PL)
        if exchange_rate > 0:
            etree.SubElement(fa, "KursWalutyZ").text = f"{exchange_rate:.4f}"

        # Sumy (EUR)
        # Logika KSeF: P_13_x / P_14_x w walucie faktury
        if is_np:
            # Dla np zwykle nie mapuje się do P_13/14, chyba ze P_13_5? 
            # Ale bezpieczniej zostawić 'np' w wierszach i P_15 jako sume.
            # Jeśli są inne stawki, trzeba dodać logikę. Na razie obsługa prosta.
            pass
        elif is_exempt:
            etree.SubElement(fa, "P_13_7").text = f"{total_netto_eur:.2f}"
        else:
            # Zakładamy stawkę podstawową dla reszty
            etree.SubElement(fa, "P_13_1").text = f"{total_netto_eur:.2f}"
            etree.SubElement(fa, "P_14_1").text = f"{total_vat_eur:.2f}"
            
        etree.SubElement(fa, "P_15").text = f"{(total_netto_eur + total_vat_eur):.2f}"

        # Adnotacje
        adnot = etree.SubElement(fa, "Adnotacje")
        for tag in ["P_16", "P_17", "P_18", "P_18A"]:
            etree.SubElement(adnot, tag).text = "2"
            
        if is_exempt:
            zw = etree.SubElement(adnot, "Zwolnienie")
            etree.SubElement(zw, "P_19").text = "1"
            etree.SubElement(zw, "P_19A").text = self.rules['vat_rules']['default_exempt_basis']
        else:
            zw = etree.SubElement(adnot, "Zwolnienie")
            etree.SubElement(zw, "P_19N").text = "1"

        nst = etree.SubElement(adnot, "NoweSrodkiTransportu")
        etree.SubElement(nst, "P_22N").text = "1"
        etree.SubElement(adnot, "P_23").text = "2"
        pm = etree.SubElement(adnot, "PMarzy")
        etree.SubElement(pm, "P_PMarzyN").text = "1"

        etree.SubElement(fa, "RodzajFaktury").text = "VAT"

        # DodatkowyOpis (PLN Info)
        idx_desc = 1
        # 1. Wartość netto PLN dla każdej pozycji
        for item in processed_items:
            dop = etree.SubElement(fa, "DodatkowyOpis")
            etree.SubElement(dop, "NrWiersza").text = str(idx_desc)
            etree.SubElement(dop, "Klucz").text = "Wartość netto w PLN"
            etree.SubElement(dop, "Wartosc").text = f"{item['PLN_Netto']:.2f}".replace('.', ',')
            idx_desc += 1
            
        # 2. Kurs
        if exchange_rate > 0:
            dop = etree.SubElement(fa, "DodatkowyOpis")
            etree.SubElement(dop, "NrWiersza").text = "1"
            etree.SubElement(dop, "Klucz").text = "Kurs PLN/EUR"
            etree.SubElement(dop, "Wartosc").text = f"{exchange_rate:.4f}".replace('.', ',')

        # Wiersze
        for idx, item in enumerate(processed_items, 1):
            w = etree.SubElement(fa, "FaWiersz")
            etree.SubElement(w, "NrWierszaFa").text = str(idx)
            etree.SubElement(w, "P_7").text = item["P_7"]
            etree.SubElement(w, "P_8A").text = item["P_8A"]
            etree.SubElement(w, "P_8B").text = item["P_8B"]
            etree.SubElement(w, "P_9A").text = item["P_9A"]
            etree.SubElement(w, "P_11").text = item["P_11"]
            if item["P_11Vat"]:
                etree.SubElement(w, "P_11Vat").text = item["P_11Vat"]
            etree.SubElement(w, "P_12").text = item["P_12"]

        # Platnosc
        plat = etree.SubElement(fa, "Platnosc")
        term = etree.SubElement(plat, "TerminPlatnosci")
        desc = etree.SubElement(term, "TerminOpis")
        etree.SubElement(desc, "Ilosc").text = "10"
        etree.SubElement(desc, "Jednostka").text = "dni/days"
        etree.SubElement(desc, "ZdarzeniePoczatkowe").text = str(data['payment_method'])
        
        etree.SubElement(plat, "FormaPlatnosci").text = "6"
        rb = etree.SubElement(plat, "RachunekBankowy")
        etree.SubElement(rb, "NrRB").text = str(data['bank_account']).replace(' ', '')
        etree.SubElement(rb, "NazwaBanku").text = p1_cfg['Nazwa']

        # Stopka
        stopka = etree.SubElement(root, "Stopka")
        inf = etree.SubElement(stopka, "Informacje")
        etree.SubElement(inf, "StopkaFaktury").text = p1_cfg.get('StopkaFaktury', '')
        rej = etree.SubElement(stopka, "Rejestry")
        etree.SubElement(rej, "PelnaNazwa").text = p1_cfg['Nazwa']
        etree.SubElement(rej, "KRS").text = "0000245000"

        # Zapisz
        tree = etree.ElementTree(root)
        tree.write(output_path, encoding="utf-8", xml_declaration=True, pretty_print=True)
        return output_path
