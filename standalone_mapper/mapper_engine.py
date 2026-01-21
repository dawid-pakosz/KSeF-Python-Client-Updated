import os
import re
import json
from datetime import datetime
from lxml import etree
import openpyxl
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string

class StandaloneMapper:
    def __init__(self, settings_path, mapping_path):
        with open(settings_path, 'r', encoding='utf-8') as f:
            self.settings = json.load(f)
        with open(mapping_path, 'r', encoding='utf-8') as f:
            self.mapping = json.load(f)
            
        self.ns = {None: "http://crd.gov.pl/wzor/2025/06/25/13775/"}

    def survey_excel(self, wb):
        """Wypisuje wszystkie znalezione nazwane zakresy dla użytkownika."""
        print("\n--- ZWIAD NAZWANYCH ZAKRESÓW ---")
        found = False
        # W nowszych wersjach openpyxl defined_names zachowuje się jak słownik
        for name in wb.defined_names.keys():
            print(f"[*] Znaleziono zakres: '{name}'")
            found = True
        if not found:
            print("[!] Nie znaleziono żadnych nazwanych zakresów w tym pliku!")
        print("--------------------------------\n")

    def get_value_by_name(self, wb, name):
        """Pobiera wartość z nazwanego zakresu (globalnego lub lokalnego)."""
        d_name = wb.defined_names.get(name)
        if not d_name:
            return None
        
        dest = list(d_name.destinations)[0] # Arkusz, Komórka
        sheet = wb[dest[0]]
        coord = dest[1].replace('$', '')
        return sheet[coord].value

    def parse_recipient_info(self, raw_text):
        """Rozbija multiline text na NIP, Nazwę i Adres."""
        if not raw_text: return {"NIP": "", "Nazwa": "", "AdresL1": ""}
        lines = [line.strip() for line in raw_text.split('\n') if line.strip()]
        result = {"NIP": "", "Nazwa": "", "AdresL1": ""}
        
        nip_match = re.search(r'NIP:\s*(\d{10})', raw_text)
        if nip_match:
            result["NIP"] = nip_match.group(1)
            lines = [l for l in lines if "NIP:" not in l]

        if lines:
            result["AdresL1"] = lines[-1]
            if len(lines) > 1:
                result["AdresL1"] = f"{lines[-2]}, {lines[-1]}"
                lines = lines[:-2]
            else:
                lines = lines[:-1]

        result["Nazwa"] = " ".join(lines)
        return result

    def process_dynamic_rows(self, sheet, start_row, start_col):
        """Skanuje tabelę w dół, rozpoznając zmiany nagłówków."""
        items = []
        col_map = {}
        curr_row = start_row

        while curr_row <= sheet.max_row:
            row_cells = list(sheet.iter_rows(min_row=curr_row, max_row=curr_row))[0]
            raw_vals = [str(c.value).strip().lower() if c.value is not None else "" for c in row_cells]
            
            # Detekcja Nagłówka
            if "nazwa usługi" in raw_vals or "nazwa uslugi" in raw_vals:
                col_map = {"P_7": raw_vals.index("nazwa usługi")}
                for tag, aliases in {
                    "P_8A": ["jm", "jednostka miary"],
                    "P_8B": ["ilość", "ilosc"],
                    "P_11": ["wartość sprzedaży netto w pln", "netto pln"],
                    "P_12": ["stawka %", "stawka vat"],
                    "cena_eur": ["cena jednostkowa netto eur", "cena eur"],
                    "kwota_vat": ["kwota w pln", "kwota vat pln"]
                }.items():
                    for a in aliases:
                        if a in raw_vals:
                            col_map[tag] = raw_vals.index(a)
                            break
                curr_row += 1
                continue

            # Dane (jeśli Lp jest cyfrą)
            val_lp = raw_vals[start_col]
            if val_lp.replace('.', '').isdigit():
                if col_map:
                    row_data = {tag: row_cells[idx].value for tag, idx in col_map.items()}
                    items.append(row_data)
            
            # Stop
            if "suma" in raw_vals or "razem" in raw_vals: break
            curr_row += 1
            
        return items

    def create_xml(self, excel_path):
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        self.survey_excel(wb)
        sheet = wb.active
        
        rules = self.mapping['mapping']
        firm = self.settings['sprzedawca']
        
        root = etree.Element("Faktura", nsmap=self.ns)
        
        # 1. Nagłówek
        nag = etree.SubElement(root, "Naglowek")
        etree.SubElement(nag, "KodFormularza", kodSystemowy="FA (3)", wersjaSchemy="1-0E").text = "FA"
        etree.SubElement(nag, "WariantFormularza").text = "3"
        etree.SubElement(nag, "DataWytworzeniaFa").text = datetime.now().isoformat() + "Z"
        etree.SubElement(nag, "SystemInfo").text = "KSeF-Mapper-Standalone-V1"
        
        # 2. Sprzedawca
        p1 = etree.SubElement(root, "Podmiot1")
        d_id1 = etree.SubElement(p1, "DaneIdentyfikacyjne")
        etree.SubElement(d_id1, "NIP").text = firm['nip']
        etree.SubElement(d_id1, "Nazwa").text = firm['nazwa']
        addr1 = etree.SubElement(p1, "Adres")
        etree.SubElement(addr1, "KodKraju").text = "PL"
        etree.SubElement(addr1, "AdresL1").text = firm['adres']
        
        # 3. Nabywca
        raw_buyer = self.get_value_by_name(wb, rules['buyer']['FullBlock']['value'])
        b_data = self.parse_recipient_info(raw_buyer)
        
        p2 = etree.SubElement(root, "Podmiot2")
        d_id2 = etree.SubElement(p2, "DaneIdentyfikacyjne")
        etree.SubElement(d_id2, "NIP").text = b_data['NIP']
        etree.SubElement(d_id2, "Nazwa").text = b_data['Nazwa']
        addr2 = etree.SubElement(p2, "Adres")
        etree.SubElement(addr2, "KodKraju").text = "PL"
        etree.SubElement(addr2, "AdresL1").text = b_data['AdresL1']
        
        # 4. Sekcja Fa
        fa = etree.SubElement(root, "Fa")
        etree.SubElement(fa, "KodWaluty").text = rules['header']['KodWaluty']['value']
        etree.SubElement(fa, "P_1").text = str(self.get_value_by_name(wb, rules['header']['P_1']['value']))
        etree.SubElement(fa, "P_2").text = str(self.get_value_by_name(wb, rules['header']['P_2']['value']))
        if 'P_6' in rules['header']:
            p6 = self.get_value_by_name(wb, rules['header']['P_6']['value'])
            if p6: etree.SubElement(fa, "P_6").text = str(p6)

        # 5. Wiersze i Sumy
        total_netto = 0
        total_vat = 0
        
        anchor = wb.defined_names.get(rules['items']['anchor_range'])
        if anchor:
            dest = list(anchor.destinations)[0]
            ws = wb[dest[0]]
            coord = dest[1].replace('$', '')
            s_row = int(re.search(r'\d+', coord).group())
            s_col = column_index_from_string(re.search(r'[A-Z]+', coord).group())
            
            items = self.process_dynamic_rows(ws, s_row - 1, s_col - 1)
            for idx, item in enumerate(items, 1):
                w = etree.SubElement(fa, "FaWiersz")
                etree.SubElement(w, "NrWierszaFa").text = str(idx)
                etree.SubElement(w, "P_7").text = str(item.get("P_7", ""))
                etree.SubElement(w, "P_8A").text = str(item.get("P_8A", ""))
                etree.SubElement(w, "P_8B").text = str(item.get("P_8B", ""))
                
                net = float(item.get("P_11") or 0)
                vat = float(item.get("kwota_vat") or 0)
                total_netto += net
                total_vat += vat
                
                etree.SubElement(w, "P_11").text = f"{net:.2f}"
                etree.SubElement(w, "P_12").text = str(item.get("P_12", "23")).replace('%', '').strip()
                
                if item.get("cena_eur"):
                    do = etree.SubElement(fa, "DodatkowyOpis")
                    etree.SubElement(do, "NrWiersza").text = str(idx)
                    etree.SubElement(do, "Klucz").text = "Cena jednostkowa netto w EUR"
                    etree.SubElement(do, "Wartosc").text = f"{float(item['cena_eur']):.2f}"

        # Sumowanie końcowe
        etree.SubElement(fa, "P_13_1").text = f"{total_netto:.2f}"
        etree.SubElement(fa, "P_14_1").text = f"{total_vat:.2f}"
        etree.SubElement(fa, "P_15").text = f"{(total_netto + total_vat):.2f}"
        
        adnot = etree.SubElement(fa, "Adnotacje")
        for tag in ["P_16", "P_17", "P_18", "P_18A", "P_23"]:
            etree.SubElement(adnot, tag).text = "2"
        etree.SubElement(fa, "RodzajFaktury").text = "VAT"

        xml_path = excel_path.replace(".xlsx", ".xml")
        tree = etree.ElementTree(root)
        tree.write(xml_path, encoding="utf-8", xml_declaration=True, pretty_print=True)
        return xml_path
